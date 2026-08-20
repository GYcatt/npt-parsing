#!/usr/bin/env python3
"""將 NPT result/<Project>/ CSV 填入 data analysis.xlsx（LTE MAX / AVG / Mid AVG）。"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

APP_DIR = Path(__file__).resolve().parent
ROOT = APP_DIR.parent
NPT_DIR = ROOT / "NPT result"
WORKBOOK = ROOT / "data analysis.xlsx"
TEMPLATE = APP_DIR / "data analysis_temp.xlsx"

SHEET_MAX = "LTE MAX"
SHEET_AVG = "LTE AVG"
SHEET_MID = "LTE Mid AVG"
SHEET_LEGACY = "LTE"
SHEET_5G = "5G"
LTE_SHEETS = (SHEET_MAX, SHEET_AVG, SHEET_MID)

STAT_MAX = "max"
STAT_AVG = "avg"
STAT_MID = "mid_avg"
SHEET_STAT = {
    SHEET_MAX: STAT_MAX,
    SHEET_AVG: STAT_AVG,
    SHEET_MID: STAT_MID,
}

PROJECT_NAME_ROW = 9
PHOTO_ROW = 8
DATA_FIRST_ROW = 10
FIRST_PROJECT_COL = 6
PAIR_WIDTH = 2

BAND_RE = re.compile(r"LTE\s+(\d+)", re.IGNORECASE)
BW_RE = re.compile(r"(\d+)\s*MHz", re.IGNORECASE)
CSV_PATTERN = re.compile(r"Band(\d+)_(\d+)MHz", re.IGNORECASE)

RX_MAIN = 0
RX_AUX = 1

FONT_PASS = "000000"
FONT_WARN = "FFC000"
FONT_FAIL = "FF0000"
FILL_EMPTY = PatternFill()

FREQ_COLUMNS = (
    "Frequency",
    "Freq",
    "RxFrequency",
    "RxFreq",
    "DL_Frequency",
    "DlFrequency",
    "DLFreq",
)


def parse_csv_name(path: Path) -> tuple[int, int] | None:
    match = CSV_PATTERN.search(path.name)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def find_freq_column(df: pd.DataFrame) -> str | None:
    lower = {str(col).strip().lower(): col for col in df.columns}
    for name in FREQ_COLUMNS:
        found = lower.get(name.lower())
        if found is not None:
            return str(found)
    return None


def load_rx_stat(
    csv_path: Path,
    rx_id: int,
    stat: str,
    bw_mhz: int | None,
    cache: dict,
) -> float | None:
    """cache keys: ('df', path), ('freq', path), (path, rx_id, stat, bw)."""
    df_key = ("df", csv_path)
    if df_key not in cache:
        cache[df_key] = pd.read_csv(csv_path)
    df = cache[df_key]
    if df.empty or "RxId" not in df.columns or "RxAGC_Value" not in df.columns:
        return None

    key = (csv_path, rx_id, stat, bw_mhz)
    if key in cache:
        return cache[key]

    subset = df[df["RxId"] == rx_id]["RxAGC_Value"].dropna()
    value: float | None = None
    if stat == STAT_MAX:
        if not subset.empty:
            value = round(float(subset.max()), 1)
    elif stat == STAT_AVG:
        if not subset.empty:
            value = round(float(subset.mean()), 1)
    elif stat == STAT_MID:
        freq_key = ("freq", csv_path)
        if freq_key not in cache:
            cache[freq_key] = find_freq_column(df)
        freq_col = cache[freq_key]
        if freq_col is not None and bw_mhz is not None and bw_mhz > 0 and freq_col in df.columns:
            freqs = df[freq_col].dropna()
            if not freqs.empty:
                mid = (float(freqs.min()) + float(freqs.max())) / 2.0
                half = bw_mhz / 2.0
                window = df[
                    (df["RxId"] == rx_id)
                    & (df[freq_col] >= mid - half)
                    & (df[freq_col] <= mid + half)
                ]["RxAGC_Value"].dropna()
                if not window.empty:
                    value = round(float(window.mean()), 1)

    cache[key] = value
    return value


def status_font_color(value: float | None, spec: float | None) -> str:
    if value is None or spec is None:
        return FONT_PASS
    if value <= spec:
        return FONT_PASS
    if value <= spec + 2:
        return FONT_WARN
    return FONT_FAIL


def apply_value_style(cell, value: float | None, spec: float | None) -> None:
    font = copy(cell.font)
    font.color = status_font_color(value, spec)
    cell.font = font


def copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def read_lte_rows(ws: Worksheet) -> dict[int, tuple[int, int | None, float | None]]:
    """band -> (row, bw MHz, spec)."""
    rows: dict[int, tuple[int, int | None, float | None]] = {}
    for row in range(DATA_FIRST_ROW, ws.max_row + 1):
        label = ws.cell(row, 2).value
        if label is None:
            continue
        match = BAND_RE.search(str(label))
        if not match:
            continue
        band = int(match.group(1))
        bw = None
        bw_raw = ws.cell(row, 4).value
        if bw_raw is not None:
            bw_match = BW_RE.search(str(bw_raw))
            if bw_match:
                bw = int(bw_match.group(1))
        spec_raw = ws.cell(row, 5).value
        spec = None
        if spec_raw is not None and spec_raw != "":
            spec = float(spec_raw)
        rows[band] = (row, bw, spec)
    return rows


def iter_project_pairs(ws: Worksheet) -> list[tuple[str, int]]:
    pairs: list[tuple[str, int]] = []
    col = FIRST_PROJECT_COL
    while col + 1 <= max(ws.max_column, FIRST_PROJECT_COL):
        name = ws.cell(PROJECT_NAME_ROW, col).value
        if name is None or not str(name).strip():
            break
        pairs.append((str(name).strip(), col))
        col += PAIR_WIDTH
    return pairs


def next_pair_col(ws: Worksheet) -> int:
    pairs = iter_project_pairs(ws)
    if not pairs:
        return FIRST_PROJECT_COL
    return pairs[-1][1] + PAIR_WIDTH


def is_preformatted_pair(ws: Worksheet, start_col: int) -> bool:
    """範本已備好表頭（第 1 列有 Main）的欄組，沿用原格式即可。"""
    return ws.cell(1, start_col).value is not None


def unmerge_cols(ws: Worksheet, start_col: int, end_col: int) -> None:
    to_remove = [
        str(merged)
        for merged in ws.merged_cells.ranges
        if merged.min_col <= end_col and merged.max_col >= start_col
    ]
    for ref in to_remove:
        ws.unmerge_cells(ref)


def copy_column_width(ws: Worksheet, src_col: int, dst_col: int) -> None:
    src = get_column_letter(src_col)
    dst = get_column_letter(dst_col)
    width = ws.column_dimensions[src].width
    if width is not None:
        ws.column_dimensions[dst].width = width


def previous_pair_col(ws: Worksheet, start_col: int) -> int:
    return max(FIRST_PROJECT_COL, start_col - PAIR_WIDTH)


def last_data_row(lte_rows: dict[int, tuple[int, int | None, float | None]]) -> int:
    if not lte_rows:
        return DATA_FIRST_ROW
    return max(row for row, _bw, _spec in lte_rows.values())


def is_merged(ws: Worksheet, row: int, start_col: int, end_col: int) -> bool:
    target = (row, start_col, row, end_col)
    return any(
        (m.min_row, m.min_col, m.max_row, m.max_col) == target
        for m in ws.merged_cells.ranges
    )


def ensure_merge(ws: Worksheet, row: int, start_col: int, end_col: int) -> None:
    if is_merged(ws, row, start_col, end_col):
        return
    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )


def clone_pair_format(ws: Worksheet, start_col: int, end_row: int) -> None:
    """表頭尚未備好時，從左邊一組複製格式（不抄 UMTS/GPS 打勾與數值）。"""
    template_col = previous_pair_col(ws, start_col)
    if template_col == start_col:
        return

    unmerge_cols(ws, start_col, start_col + 1)
    for offset in range(PAIR_WIDTH):
        copy_column_width(ws, template_col + offset, start_col + offset)
        for row in range(1, end_row + 1):
            src = ws.cell(row, template_col + offset)
            dst = ws.cell(row, start_col + offset)
            copy_cell_style(src, dst)
            if row == 1:
                dst.value = "Main" if offset == 0 else "Aux"
                continue
            dst.value = None
            font = copy(dst.font)
            font.color = FONT_PASS
            dst.font = font
            if row >= DATA_FIRST_ROW:
                dst.fill = FILL_EMPTY


def setup_project_pair(ws: Worksheet, start_col: int, project_name: str, lte_rows: dict) -> None:
    end_row = max(PROJECT_NAME_ROW, last_data_row(lte_rows))

    if not is_preformatted_pair(ws, start_col):
        clone_pair_format(ws, start_col, end_row)

    ensure_merge(ws, PHOTO_ROW, start_col, start_col + 1)
    ensure_merge(ws, PROJECT_NAME_ROW, start_col, start_col + 1)
    ws.cell(PROJECT_NAME_ROW, start_col, project_name)


def pick_csv_files(project_dir: Path) -> dict[tuple[int, int], Path]:
    """(band, bw) -> latest CSV."""
    selected: dict[tuple[int, int], Path] = {}
    for csv_path in sorted(project_dir.glob("*.csv")):
        parsed = parse_csv_name(csv_path)
        if not parsed:
            continue
        band, bw = parsed
        key = (band, bw)
        prev = selected.get(key)
        if prev is None or csv_path.name > prev.name:
            selected[key] = csv_path
    return selected


def write_project_values(
    ws: Worksheet,
    start_col: int,
    csv_map: dict[tuple[int, int], Path],
    lte_rows: dict[int, tuple[int, int | None, float | None]],
    stat: str,
    cache: dict,
) -> tuple[list[int], list[int]]:
    missing: list[int] = []
    mid_empty: list[int] = []

    for band, (row, bw, spec) in sorted(lte_rows.items()):
        csv_path = csv_map.get((band, bw)) if bw is not None else None
        main_val = aux_val = None
        if csv_path is None:
            missing.append(band)
        else:
            main_val = load_rx_stat(csv_path, RX_MAIN, stat, bw, cache)
            aux_val = load_rx_stat(csv_path, RX_AUX, stat, bw, cache)
            if stat == STAT_MID and main_val is None and aux_val is None:
                mid_empty.append(band)

        main_cell = ws.cell(row, start_col)
        aux_cell = ws.cell(row, start_col + 1)
        main_cell.value = main_val
        aux_cell.value = aux_val
        apply_value_style(main_cell, main_val, spec)
        apply_value_style(aux_cell, aux_val, spec)

    return missing, mid_empty


def clear_project_values(ws: Worksheet, lte_rows: dict) -> None:
    """清空專案數字，保留第 9 列名稱與表頭格式。"""
    pairs = [col for _name, col in iter_project_pairs(ws)]
    end_row = last_data_row(lte_rows)
    for start_col in pairs:
        for offset in range(PAIR_WIDTH):
            for row in range(DATA_FIRST_ROW, end_row + 1):
                cell = ws.cell(row, start_col + offset)
                cell.value = None
                font = copy(cell.font)
                font.color = FONT_PASS
                cell.font = font


def reset_project_columns(ws: Worksheet, lte_rows: dict) -> int:
    """清空所有已命名的 Project 欄（保留表頭格式），回傳清除組數。"""
    pairs = [col for _name, col in iter_project_pairs(ws)]
    clear_project_values(ws, lte_rows)
    for start_col in pairs:
        ws.cell(PROJECT_NAME_ROW, start_col).value = None
    return len(pairs)


def strip_draft(ws: Worksheet, lte_rows: dict) -> None:
    """刪除主表 Band 列以下的「本次新增資料」區塊。"""
    end = last_data_row(lte_rows)
    if ws.max_row <= end:
        return
    to_remove = [
        str(merged)
        for merged in ws.merged_cells.ranges
        if merged.min_row > end
    ]
    for ref in to_remove:
        ws.unmerge_cells(ref)
    ws.delete_rows(end + 1, ws.max_row - end)


def order_sheets(wb) -> None:
    wanted = [SHEET_MAX, SHEET_AVG, SHEET_MID, SHEET_5G]
    existing = [name for name in wanted if name in wb.sheetnames]
    rest = [name for name in wb.sheetnames if name not in existing]
    wb._sheets = [wb[name] for name in existing + rest]


def clone_avg_sheet(wb, src_name: str, dst_name: str, lte_rows: dict) -> None:
    dst = wb.copy_worksheet(wb[src_name])
    dst.title = dst_name
    strip_draft(dst, lte_rows)
    clear_project_values(dst, lte_rows)


def prepare_lte_sheets(wb) -> list[str]:
    """舊 LTE 改名、補平均表、調整分頁順序。回傳訊息。"""
    notes: list[str] = []
    names = set(wb.sheetnames)

    if SHEET_LEGACY in names and SHEET_MAX not in names:
        wb[SHEET_LEGACY].title = SHEET_MAX
        notes.append(f"已將「{SHEET_LEGACY}」改名為「{SHEET_MAX}」")
        names = set(wb.sheetnames)
    elif SHEET_LEGACY in names and SHEET_MAX in names:
        notes.append(f"同時存在「{SHEET_LEGACY}」與「{SHEET_MAX}」，以「{SHEET_MAX}」為準")

    if SHEET_MAX not in wb.sheetnames:
        return notes

    lte_rows = read_lte_rows(wb[SHEET_MAX])
    for sheet_name in (SHEET_AVG, SHEET_MID):
        if sheet_name in wb.sheetnames:
            continue
        clone_avg_sheet(wb, SHEET_MAX, sheet_name, lte_rows)
        notes.append(f"已新增「{sheet_name}」（自 {SHEET_MAX} 複製版面）")

    order_sheets(wb)
    return notes


def ensure_workbook() -> bool:
    """結果檔不存在時，從 app/data analysis_temp.xlsx 複製一份。"""
    if WORKBOOK.exists():
        return True
    if not TEMPLATE.exists():
        print(f"找不到結果檔，也找不到範本：{TEMPLATE}")
        return False
    shutil.copy2(TEMPLATE, WORKBOOK)
    print(f"找不到 {WORKBOOK.name}，已從範本 {TEMPLATE.name} 建立新檔。")
    return True


def process_projects(rebuild: bool = False) -> int:
    if not ensure_workbook():
        return 1
    if not NPT_DIR.is_dir():
        print(f"找不到資料夾：{NPT_DIR}")
        return 1

    wb = load_workbook(WORKBOOK)
    for line in prepare_lte_sheets(wb):
        print(line)

    if SHEET_MAX not in wb.sheetnames:
        print(f"找不到 {SHEET_MAX} sheet")
        return 1

    sheets = {name: wb[name] for name in LTE_SHEETS if name in wb.sheetnames}
    ws_max = sheets[SHEET_MAX]
    lte_rows = read_lte_rows(ws_max)
    if not lte_rows:
        print(f"{SHEET_MAX} 找不到 Band 列。")
        return 1

    if rebuild:
        removed = 0
        for ws in sheets.values():
            removed = reset_project_columns(ws, lte_rows)
        print(f"強制重算：已清空 {removed} 組 Project 欄，依 CSV 全量重寫…")

    existing = {name: col for name, col in iter_project_pairs(ws_max)}
    project_dirs = sorted(p for p in NPT_DIR.iterdir() if p.is_dir())
    if not project_dirs:
        print("NPT result/ 內沒有 Project 資料夾。")
        try:
            wb.save(WORKBOOK)
        except PermissionError:
            print(f"無法寫入 {WORKBOOK.name}，請先關閉 Excel 再重新執行。")
            return 1
        return 0

    added = 0
    updated = 0
    freq_warned = False
    for project_dir in project_dirs:
        name = project_dir.name
        csv_map = pick_csv_files(project_dir)
        if not csv_map:
            print(f"跳過（無符合 CSV）：{name}")
            continue

        cache: dict = {}
        start_col = existing.get(name)
        if start_col is None:
            start_col = next_pair_col(ws_max)
            for ws in sheets.values():
                setup_project_pair(ws, start_col, name, lte_rows)
            existing[name] = start_col
            added += 1
            action = "已追加"
        else:
            for ws in sheets.values():
                setup_project_pair(ws, start_col, name, lte_rows)
            updated += 1
            action = "已覆寫"

        missing: list[int] = []
        mid_empty: list[int] = []
        for sheet_name, ws in sheets.items():
            miss, mid_miss = write_project_values(
                ws, start_col, csv_map, lte_rows, SHEET_STAT[sheet_name], cache
            )
            if sheet_name == SHEET_MAX:
                missing = miss
            if sheet_name == SHEET_MID:
                mid_empty = mid_miss

        if not freq_warned and SHEET_MID in sheets:
            sample = next(iter(csv_map.values()), None)
            if sample is not None:
                df = cache.get(("df", sample))
                if df is not None and find_freq_column(df) is None:
                    print("警告：CSV 找不到頻率欄，LTE Mid AVG 全部留空")
                    freq_warned = True

        miss_txt = f"，缺檔 Band：{', '.join(f'B{b}' for b in missing)}" if missing else ""
        mid_txt = (
            f"，Mid 視窗無資料：{', '.join(f'B{b}' for b in mid_empty)}"
            if mid_empty and not freq_warned
            else ""
        )
        print(
            f"{action}：{name}（{len(csv_map)} 個 CSV）→ 欄 {get_column_letter(start_col)}"
            f"{miss_txt}{mid_txt}"
        )

    try:
        wb.save(WORKBOOK)
    except PermissionError:
        print(f"無法寫入 {WORKBOOK.name}，請先關閉 Excel 再重新執行。")
        return 1

    print(f"完成，追加 {added} 組，覆寫 {updated} 組。")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="NPT CSV → data analysis.xlsx LTE MAX / AVG / Mid AVG"
    )
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="清空所有 Project 欄後，依 NPT result 全量重寫",
    )
    args = parser.parse_args(argv)
    return process_projects(rebuild=args.rebuild)


if __name__ == "__main__":
    sys.exit(main())
